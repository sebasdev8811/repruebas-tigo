from fastapi import APIRouter, Depends, HTTPException, status, Header
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from datetime import date
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.database import get_db
from app.models import Reprueba, DañoPendiente, GestionAnalista, GestionGarantia, Cliente, DañoCerrado, Usuario
from app.core.security import decode_token
from sqlalchemy import func

router = APIRouter(prefix="/exportes", tags=["Exportes"])

AZUL = "003D82"
AZUL_CLARO = "D5E8F0"
BLANCO = "FFFFFF"

def get_current_user(authorization: str = Header(None), db: Session = Depends(get_db)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Token no proporcionado")
    token = authorization.split(" ", 1)[1]
    payload = decode_token(token)
    if not payload or "sub" not in payload:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Token inválido o expirado")
    usuario = db.query(Usuario).filter(Usuario.id == int(payload["sub"])).first()
    if not usuario or not usuario.activo:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Usuario no válido")
    return usuario

def estilo_header(cell):
    cell.font = Font(bold=True, color=BLANCO, name="Arial", size=11)
    cell.fill = PatternFill("solid", fgColor=AZUL)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    lado = Side(style="thin", color="CCCCCC")
    cell.border = Border(top=lado, bottom=lado, left=lado, right=lado)

def estilo_celda(cell, fila_par=False):
    cell.font = Font(name="Arial", size=10)
    cell.fill = PatternFill("solid", fgColor=AZUL_CLARO if fila_par else BLANCO)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    lado = Side(style="thin", color="CCCCCC")
    cell.border = Border(top=lado, bottom=lado, left=lado, right=lado)

def crear_excel(titulo, headers, filas):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo[:31]
    ws.row_dimensions[1].height = 30

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        estilo_header(cell)
        ws.column_dimensions[cell.column_letter].width = max(len(header) + 4, 15)

    for row_idx, fila in enumerate(filas, 2):
        par = row_idx % 2 == 0
        for col_idx, valor in enumerate(fila, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            estilo_celda(cell, par)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.get("/repruebas")
async def exportar_repruebas(
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    query = db.query(Reprueba, DañoPendiente, DañoCerrado).outerjoin(
        DañoPendiente, Reprueba.pedido_id == DañoPendiente.pedido_id
    ).outerjoin(
        DañoCerrado, Reprueba.pedido_id == DañoCerrado.pedido_id
    )
    if fecha_inicio:
        query = query.filter(Reprueba.fecha_reprueba >= fecha_inicio)
    if fecha_fin:
        query = query.filter(Reprueba.fecha_reprueba <= fecha_fin)

    registros = query.all()

    headers = ["ID Reprueba", "Pedido ID", "Código Estado", "Fecha Reprueba", "Tipo Falla", "Ciudad", "Nodo", "CMTS", "Amplificador", "TAP"]
    filas = []
    for rep, dano_p, dano_c in registros:
        dano = dano_p or dano_c
        filas.append([
            rep.id_reprueba,
            rep.pedido_id,
            rep.codigo_estado,
            str(rep.fecha_reprueba) if rep.fecha_reprueba else "",
            dano.tipo_falla if dano else "",
            dano.ciudad if dano else "",
            dano.nodo if dano else "",
            dano.cmts if dano else "",
            dano.amplificador if dano else "",
            dano.tap if dano else "",
        ])

    output = crear_excel("Repruebas", headers, filas)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=repruebas.xlsx"}
    )


@router.get("/fallas-masivas")
async def exportar_fallas_masivas(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    gestiones = db.query(GestionAnalista).all()

    headers = ["ID", "Elemento Red", "Valor", "Total Pedidos", "Estado", "Observaciones", "Fecha Detección", "Fecha Gestión", "Ticket"]
    filas = []
    for g in gestiones:
        filas.append([
            g.id,
            g.elemento_red,
            g.valor_elemento,
            g.total_pedidos,
            g.estado,
            g.observaciones or "",
            str(g.fecha_deteccion)[:10] if g.fecha_deteccion else "",
            str(g.fecha_gestion)[:10] if g.fecha_gestion else "",
            g.numero_ticket or "",
        ])

    output = crear_excel("Fallas Masivas", headers, filas)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=fallas_masivas.xlsx"}
    )


@router.get("/garantias")
async def exportar_garantias(
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    query = db.query(GestionGarantia, DañoCerrado, Cliente).outerjoin(
        DañoCerrado, GestionGarantia.pedido_id == DañoCerrado.pedido_id
    ).outerjoin(
        Cliente, DañoCerrado.cedula_cliente == Cliente.cedula_cliente
    )
    if fecha_inicio and fecha_fin:
        query = query.filter(
            func.date(GestionGarantia.fecha_gestion) >= fecha_inicio,
            func.date(GestionGarantia.fecha_gestion) <= fecha_fin
        )

    registros = query.all()

    headers = ["ID", "Pedido ID", "Cliente", "Cédula", "Ciudad", "Tipo Falla", "Estado Reprueba", "Estado Gestión", "Observaciones", "Fecha Gestión", "Ticket"]
    filas = []
    for g, dano, cliente in registros:
        filas.append([
            g.id,
            g.pedido_id,
            cliente.nombre if cliente else "",
            cliente.cedula_cliente if cliente else "",
            dano.ciudad if dano else "",
            dano.tipo_falla if dano else "",
            g.codigo_estado_reprueba or "",
            g.estado,
            g.observaciones or "",
            str(g.fecha_gestion)[:10] if g.fecha_gestion else "",
            g.numero_ticket or "",
        ])

    output = crear_excel("Garantias", headers, filas)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=garantias.xlsx"}
    )